import os
import json
import codecs
import numpy as np
import pandas as pd

from openpyxl import load_workbook

from Method.SensorData import Data
from Algorithm.FNN import FNN


class LoadData:
    # It's the Data path for the google drive
    # DATA_DIR = "C:\\Users\\User\\Google 雲端硬碟\\畢業專題\\安全帽\\LabelingData"
    # file = load_workbook(os.path.join(DATA_DIR, '074.xlsx'))

    def __init__(self):
        # It's the column in the xlsx
        self.data_list = []
        self.Acc = []
        self.Gyro = []
        self.Mag = []
        self.Pre = []
        self.sensor_dim = []

    # load Data from excel
    def get_stsensor_excel(self, file):
        # It's the Data path for the local path in yu ren computer
        rel_data_dir = '../Data/LabelingData/' + file + '.xlsx'
        abs_data_dir = os.path.join(os.path.dirname(__file__), rel_data_dir)
        file = load_workbook(abs_data_dir)
        xlsx = file['STSensor']
        for row in range(2, xlsx.max_row + 1):
            date, time, hall = '', '', ''
            for col in range(1, xlsx.max_column + 1):
                data = str(xlsx.cell(row=row, column=col).value)
                if col == 1:
                    date = str(data)
                if col == 2:
                    time = str(data)
                if 3 <= col <= 5:
                    self.Acc.append(data)
                if 6 <= col <= 8:
                    self.Gyro.append(data)
                if 9 <= col <= 11:
                    self.Mag.append(data)
                if 12 <= col <= 13:
                    self.Pre.append(data)
                if col == 14:
                    hall = str(data)

            # Sensor Data in the xlsx
            self.data_list.append(Data(date, time, self.Acc, self.Gyro, self.Mag, self.Pre, hall))
            self.Acc, self.Gyro, self.Mag, self.Pre = [], [], [], []
        # for line in self.data_list:
        #     print(line.Date, line.time, line.Acc, line.Gyro, line.Mag, line.Pre, line.Hall)
        return self.data_list

    # @staticmethod
    # def get_analysis_excel(file):
    #     rel_data_dir = '../Data/LabelingData/' + file + '.xlsx'
    #     abs_data_dir = os.path.join(os.path.dirname(__file__), rel_data_dir)
    #     file = load_workbook(abs_data_dir)
    #     xlsx = file['分析']
    #     print(xlsx)
    #
    # def get_2dimension_data(self, file_name):
    #     path_name = '../Data/2dimension(' + file_name + ').txt'
    #     full_name = os.path.join(os.path.dirname(__file__), path_name)
    #     data, labels = [], []
    #     with open(full_name, 'r') as f:
    #         tmp = []
    #         while True:
    #             dd = f.readline()
    #             if dd == '':
    #                 break
    #             tmp.append(dd)
    #         for e in tmp:
    #             tt = e.split(';')
    #             data.append(tt[0])
    #             data.append(tt[1])
    #             labels.append(tt[2])
    #         np_data = np.array(data)
    #         np_data = np_data.astype('float64')
    #         np_labels = np.array(labels)
    #         np_labels = np_labels.astype('int64')
    #         # type -> float32
    #         # not type -> float64
    #     f.close()
    #     return np_data, np_labels

    # Load the Labeling from excel
    def get_original_data(self, label_type, file_name):
        df = pd.read_excel('../Data/Labeling/' + label_type + '/' + file_name + '.xlsx', index=True)
        # print(df.head())
        for i in range(len(df.index)):
            tmp = df.iloc[i:i+1, :]
            self.sensor_dim.append(Data(tmp.iat[0, 0], tmp.iat[0, 1], tmp.iat[0, 2],
                                        tmp.iat[0, 3], tmp.iat[0, 4], tmp.iat[0, 5],
                                        tmp.iat[0, 6], tmp.iat[0, 7], tmp.iat[0, 8],
                                        tmp.iat[0, 9], tmp.iat[0, 10], tmp.iat[0, 11],
                                        tmp.iat[0, 12], tmp.iat[0, 13], tmp.iat[0, 14]))
        return self.sensor_dim

    @staticmethod
    def get_method1_test():
        header = ['Dim' + str(i) for i in range(1, 265, 1)]
        path_name = '../Data/Labeling/C/Original_data.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header].values
        labels = excel_data.loc[:, ['Label']].values.ravel()
        labels_new = np.array([int(e[1]) for e in labels])
        return data, labels_new

    @staticmethod
    def get_method2_test():
        header = ['Dim' + str(i) for i in range(1, 4, 1)]
        path_name = '../Data/Labeling/C/method2/Test_data.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header].values
        labels = excel_data.loc[:, ['Label']].values.ravel()
        return data, labels

    @staticmethod
    def get_method1_fnn_train(num):
        """
        Read the FNN_Train_data as np_data.
        :param num: label_num (C1~C6)
        :return: np_data, np_label
        """
        header = ['Dim' + str(i) for i in range(1, 265, 1)]
        path_name = '../Data/Labeling/C/method1/FNN_Train_data_' + num + '.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header].values
        labels = excel_data.loc[:, ['Label']].values.ravel()
        labels_new = np.array([int(e[1]) for e in labels])
        return data, labels_new

    @staticmethod
    def get_method2_fnn_train(num):
        header = ['Dim' + str(i) for i in range(1, 4, 1)]
        path_name = '../Data/Labeling/C/method2/FNN_Train_data_' + num + '.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header].values
        labels = excel_data.loc[:, ['Label']].values.ravel()
        return data, labels

    @staticmethod
    def get_lnn_training_data():
        """Read the LNN_Train_data as np_data."""
        header = ['Dim' + str(i) for i in range(1, 265, 1)]
        path_name = '../Data/Labeling/C/LNN_Train_data.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header].values
        labels = excel_data.loc[:, ['Label']].values.ravel()
        return data, labels

    @staticmethod
    def get_org_data():
        """Read the Original_data as DataFrame."""
        header = ['Dim' + str(i) for i in range(1, 265, 1)]
        header.append('Label')
        path_name = '../Data/Labeling/C/Original_data.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header]
        return data

    @staticmethod
    def get_original_excel(label_type, file):
        """Read the 0.74.xlsx as np_array."""
        path_name = '../Data/Labeling/' + label_type + '/' + file + '.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        columns = ['AccX', 'AccY', 'AccZ',
                   'GyroX', 'GyroY', 'GyroZ',
                   'MagX', 'MagY', 'MagZ', 'PreX', 'PreY']
        data = excel_data.loc[:, columns].values
        labels = excel_data.loc[:, ['Label']].values
        return data, labels

    @staticmethod
    def get_split_data():
        """Read the Split_data as np_array."""
        path_name = '../Data/Labeling/C/Split_data.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        columns = ['AccX', 'AccY', 'AccZ',
                   'GyroX', 'GyroY', 'GyroZ',
                   'MagX', 'MagY', 'MagZ', 'PreX', 'PreY']
        data = excel_data.loc[:, columns].values
        labels = excel_data.loc[:, ['Label']].values
        return data, labels

    @staticmethod
    def get_split_original_data(num):
        """
        Read the C1_Original_data as the np_array
        :param num: label_num (C1~C6)
        :return: np_data, np_label
        """
        header = ['Dim' + str(i) for i in range(1, 265, 1)]
        # header.append('Label')
        path_name = '../Data/Labeling/C/method2/C'+str(num)+'_Original_data.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header].values
        labels = excel_data.loc[:, ['Label']].values.ravel()
        labels_new = np.array([int(e[1]) for e in labels])
        return data, labels_new

    @staticmethod
    def get_refactor_data(label, num):
        """
        Read the C1_Refactor_data.xlsx as np_array
        :param label: label_num (C1~C6)
        :param num: according the cluster num
        :return: np_data, np_label
        """
        header = ['Dim' + str(i) for i in range(1, 4, 1)]
        path_name = '../Data/Labeling/C/method2/'+label+'_'+str(num)+'_Refactor_data.xlsx'
        path_name = os.path.join(os.path.dirname(__file__), path_name)
        excel_data = pd.read_excel(path_name)
        data = excel_data.loc[:, header].values
        labels = excel_data.loc[:, ['Label']].values.ravel()
        return data, labels

    @staticmethod
    def load_fnn_weight(path):
        obj_text = codecs.open(path, 'r', encoding='utf-8').read()
        data = json.loads(obj_text)
        return data
